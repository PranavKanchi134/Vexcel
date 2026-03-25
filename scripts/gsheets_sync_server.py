#!/usr/bin/env python3
"""Local bridge used by the Vexcel popup Sync button."""

from __future__ import annotations

import json
import os
import warnings
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Dict, List, Optional, Tuple

warnings.filterwarnings("ignore", category=FutureWarning, module=r"google(\.|$)")

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from googleapiclient.discovery import build

from gsheets_cli import get_credentials


HOST = "127.0.0.1"
PORT = int(os.environ.get("VEXCEL_SYNC_PORT", "8765"))


def make_args() -> SimpleNamespace:
    root = Path(__file__).resolve().parents[1]
    return SimpleNamespace(
        service_account=os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")),
        oauth_client_secret=os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET_FILE", ""),
        token_file=os.environ.get("GOOGLE_OAUTH_TOKEN_FILE", str(root / ".secrets" / "google-sheets-token.json")),
    )


def sheets_service() -> Any:
    creds = get_credentials(make_args(), allow_interactive=False, force=False)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def find_target_sheet(spreadsheet_id: str, target_title: str) -> Dict[str, Any]:
    service = sheets_service()
    payload = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties(sheetId,title,gridProperties)"
    ).execute()
    sheets = payload.get("sheets", [])
    normalized_target = normalize_title(target_title)

    for sheet in sheets:
      props = sheet.get("properties", {})
      title = props.get("title", "")
      if normalize_title(title) == normalized_target:
          return props

    raise ValueError(f'Could not find target sheet "{target_title}" in the current spreadsheet.')


def normalize_title(value: str) -> str:
    return " ".join((value or "").replace("selected", "").split()).strip().lower()


def quote_sheet_title(title: str) -> str:
    escaped = title.replace("'", "''")
    return f"'{escaped}'"


def sync_workbook_to_sheet(spreadsheet_id: str, target_sheet_title: str, file_path: str) -> Dict[str, Any]:
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Sync currently supports .xlsx and .xlsm files.")

    workbook = load_workbook(path, data_only=False)
    source_sheet = pick_source_sheet(workbook, target_sheet_title)
    values, row_count, column_count = worksheet_values(source_sheet)

    service = sheets_service()
    target_props = find_target_sheet(spreadsheet_id, target_sheet_title)
    target_title = target_props["title"]
    sheet_id = target_props["sheetId"]
    quoted_title = quote_sheet_title(target_title)

    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=quoted_title,
        body={}
    ).execute()

    if row_count and column_count:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{quoted_title}!A1:{a1(column_count)}{row_count}",
            valueInputOption="USER_ENTERED",
            body={"majorDimension": "ROWS", "values": values}
        ).execute()

    resize_requests = [{
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "gridProperties": {
                    "rowCount": max(row_count + 50, 1000),
                    "columnCount": max(column_count + 10, 26)
                }
            },
            "fields": "gridProperties.rowCount,gridProperties.columnCount"
        }
    }]
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": resize_requests}
    ).execute()

    return {
        "ok": True,
        "spreadsheetId": spreadsheet_id,
        "sheetTitle": target_title,
        "sourceWorksheet": source_sheet.title,
        "sourceFile": str(path),
        "rowCount": row_count,
        "columnCount": column_count,
    }


def pick_source_sheet(workbook: Any, target_sheet_title: str) -> Worksheet:
    normalized_target = normalize_title(target_sheet_title)
    for sheet in workbook.worksheets:
        if normalize_title(sheet.title) == normalized_target:
            return sheet
    return workbook.active


def worksheet_values(ws: Worksheet) -> Tuple[List[List[Any]], int, int]:
    max_row = 0
    max_col = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if cell.row > max_row:
                    max_row = cell.row
                if cell.column > max_col:
                    max_col = cell.column

    if max_row == 0 or max_col == 0:
        return [], 0, 0

    rows: List[List[Any]] = []
    for row_idx in range(1, max_row + 1):
        rendered_row: List[Any] = []
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            rendered_row.append("" if value is None else value)
        rows.append(rendered_row)

    return rows, max_row, max_col


def a1(column_index: int) -> str:
    letters = ""
    current = column_index
    while current > 0:
        current, rem = divmod(current - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self) -> None:  # noqa: N802
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self) -> None:  # noqa: N802
        if self.path != "/health":
            self._json(404, {"ok": False, "error": "not_found"})
            return
        self._json(200, {"ok": True, "service": "vexcel-gsheets-sync", "port": PORT})

    def do_POST(self) -> None:  # noqa: N802
        if self.path != "/sync":
            self._json(404, {"ok": False, "error": "not_found"})
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length) or b"{}")
            result = sync_workbook_to_sheet(
                spreadsheet_id=str(payload.get("spreadsheetId", "")).strip(),
                target_sheet_title=str(payload.get("sheetTitle", "")).strip(),
                file_path=str(payload.get("filePath", "")).strip(),
            )
            self._json(200, result)
        except Exception as exc:  # pragma: no cover
            self._json(400, {"ok": False, "error": str(exc)})

    def log_message(self, format: str, *args: Any) -> None:  # noqa: A003
        return

    def _cors(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")

    def _json(self, status: int, payload: Dict[str, Any]) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self._cors()
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> int:
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"Vexcel Sheets sync server listening on http://{HOST}:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
